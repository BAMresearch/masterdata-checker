import openpyxl
import pytest

from masterdata_checker.content import (
    check_properties,
    content_checker,
    index_to_excel_column,
)


def test_index_to_excel_column():
    # Test single letters
    assert index_to_excel_column(1) == 'A'
    assert index_to_excel_column(26) == 'Z'
    # Test double letters
    assert index_to_excel_column(27) == 'AA'
    assert index_to_excel_column(52) == 'AZ'
    assert index_to_excel_column(53) == 'BA'
    # Test invalid inputs
    assert index_to_excel_column(0) == ''
    assert index_to_excel_column(-1) == ''


def test_check_properties_header_wrong():
    file_path = 'data/object_type_CHEMICAL_v1_S.3_cmadaria_error_prop_headers.xlsx'
    workbook = openpyxl.load_workbook(file_path)
    errors = []

    sheet = workbook.active

    real_errors = check_properties(sheet, errors)

    errors = []

    expected_terms = [
        'Version',
        'Code',
        'Description',
        'Mandatory',
        'Show in edit views',
        'Section',
        'Property label',
        'Data type',
        'Vocabulary code',
    ]

    for term in expected_terms:
        if term in ('Mandatory', 'Show in edit views', 'Section'):
            errors.append(f"Warning: '{term}' not found in the properties headers.")
        else:
            errors.append(f"Error: '{term}' not found in the properties headers.")

    assert real_errors == errors


def test_check_properties_header_ok():
    file_path = 'data/object_type_CHEMICAL_v1_S.3_cmadaria.xlsx'
    workbook = openpyxl.load_workbook(file_path)
    errors = []

    sheet = workbook.active

    errors = check_properties(sheet, errors)

    assert errors == []  # Should have no errors with valid data


def test_check_properties_wrong():
    file_path = 'data/object_type_CHEMICAL_v1_S.3_cmadaria_error_prop.xlsx'
    workbook = openpyxl.load_workbook(file_path)
    errors = []

    sheet = workbook.active

    errors = [
        "Error: Values not valid found in the 'Version' column (they should be Integers) at row(s): 7",
        "Error: Invalid code found in the 'Code' column at row(s): 7",
        "Error: Invalid value(s) found in the 'Description' column at row(s): 24. Description should follow the schema: English Description + '//' + German Description.",
        "Error: Invalid value found in the 'Mandatory' column at row(s): 7. Accepted values: TRUE, FALSE",
        "Error: Invalid value found in the 'Show in edit views' column at row(s): 7. Accepted values: TRUE, FALSE",
        "Error: Invalid value found in the 'Section' column at row(s): 27. Each word in the Section should start with a capital letter.",
        "Error: Invalid value found in the 'Data type' column at row(s): 23. Accepted types: INTEGER, REAL, VARCHAR, MULTILINE_VARCHAR, HYPERLINK, BOOLEAN, CONTROLLEDVOCABULARY, XML, TIMESTAMP, DATE, SAMPLE",
        "Error: Invalid vocabulary code found in the 'Vocabulary code' column at row(s): 25, 30",
    ]

    real_errors = check_properties(sheet, errors)

    assert errors == real_errors


def test_check_properties_ok():
    file_path = 'data/object_type_CHEMICAL_v1_S.3_cmadaria.xlsx'
    workbook = openpyxl.load_workbook(file_path)
    errors = []

    sheet = workbook.active

    errors = check_properties(sheet, errors)

    assert errors == []  # Should have no errors with valid data
