# -*- coding: utf-8 -*-
"""
Created on Mon Dec  4 16:46:36 2023

@author: cmadaria
"""
import openpyxl
import re

file_path = 'C:/Users/cmadaria/Documents/Projects/Type checker/object_type_CHEMICAL_v1_S.3_relathma.xlsx'
def content_checker(file_path):
    workbook = openpyxl.load_workbook(file_path)
    errors = []
    file_name = file_path.split("/")[-1]
    file_name = file_name.split(".xls")
    file_parts = file_name[0].split("_")
    file_parts.pop(-1)
    file_parts.pop(-1)
    version = file_parts.pop(-1)
    etype = file_parts.pop(0)
    if (etype == "object" or etype == "collection" or etype == "dataset"):
        etype = etype + "_" + file_parts.pop(0)
    code = "_".join(file_parts)

    # Select the specific sheet (replace 'Sheet1' with your sheet name)
    sheet = workbook.active

    # Access a specific cell (e.g., cell A1)
    cell_value_A1 = sheet['A1'].value
    print(f"Entity Type: {cell_value_A1}")
    
    entity_types = ["SAMPLE_TYPE", "EXPERIMENT_TYPE", "DATASET_TYPE", "PROPERTY_TYPE", "VOCABULARY_TYPE"]
    if cell_value_A1 not in entity_types:
        errors.append("The entity type (cell A1) should be one of the following: SAMPLE_TYPE, EXPERIMENT_TYPE, DATASET_TYPE, PROPERTY_TYPE, VOCABULARY_TYPE")
        return errors
    else:
        if cell_value_A1 == "SAMPLE_TYPE":
            expected_terms = [
                "Version",
                "Code",
                "Description",
                "Validation script",
                "Generated code prefix",
                "Auto generate codes",
            ]
            second_row_values = [cell.value for cell in sheet[2]]
            for term in expected_terms:
                if term not in second_row_values:
                    errors.append(f"Error: '{term}' not found in the second row.")
                else:
                     # Find the index of the term in the second row
                     term_index = second_row_values.index(term)

                     # Check the cell below "Version"
                     if term == "Version":
                        cell_below_version = sheet.cell(row=3, column=term_index + 1)
                        if str(cell_below_version.value) != version[1:]:
                            errors.append("Error: The version should be the same one indicated in the file name")

                    # Check the cell below "Code"
                     elif term == "Code":
                        cell_below_code = sheet.cell(row=3, column=term_index + 1)
                        if cell_below_code.value != code:
                            errors.append("Error: The code should be the same one indicated in the file name")
                    
                    
                    # Check the cell below "Description"
                     elif term == "Description":
                        cell_below_description = sheet.cell(row=3, column=term_index + 1)
                        description_pattern = re.compile(r".*//.*")
                        if not description_pattern.match(cell_below_description.value):
                            errors.append("Error: Value below 'Description' does not match the expected pattern.")

                    # Check the cell below "Generated code prefix"
                     elif term == "Generated code prefix":
                        cell_below_generated_code = sheet.cell(row=3, column=term_index + 1)
                        if cell_below_generated_code.value not in code:
                            errors.append("Error: 'code' not found in the value below 'Generated code prefix'.")

                    # Check the cell below "Auto generate codes"
                     elif term == "Auto generate codes":
                        cell_below_auto_generate = sheet.cell(row=3, column=term_index + 1)
                        if cell_below_auto_generate.value not in ["TRUE", "FALSE"]:
                            errors.append("Error: Value below 'Auto generate codes' should be 'TRUE' or 'FALSE'.")


            print(errors)


    # Close the workbook after use
    workbook.close()
    
content_checker(file_path)