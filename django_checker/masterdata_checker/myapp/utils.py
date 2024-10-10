# -*- coding: utf-8 -*-
"""
Created on Thu Oct 10

@author: cmadaria
"""

import re
import openpyxl
import logging

logger = logging.getLogger('myapp')

def name_checker(file_name):
    # Define the pattern for a valid file name
    pattern = r"^(collection_type|object_type|dataset_type|vocabulary)_([\w.]+)_(v\d+)_([a-zA-Z0-9]+(?:\.[0-9]+)?)_([a-zA-Z0-9]+)\.(xls|xlsx)$"
 
    # Check if the file name matches the pattern
    match = re.match(pattern, file_name)

    if match:
        # Extract parts of the file name
        entity_type, entity_name, version, division, contact_person, extension = match.groups()
        print(entity_type, entity_name, version, division, contact_person, extension)
        return ["<strong>File name: OK!</strong>", True]
    else:
        # Return specific errors and positions
        errors = []
        file_name = file_name.split(".xls")
        file_parts = file_name[0].split("_")
        if len(file_parts) < 5:
            errors.append("<strong>Invalid name format</strong>. The name should contain different fields separated by underscores (_). Consult the wiki to see which ones.")
            return ["\n".join(errors), False]
        creator = file_parts.pop(-1)
        section = file_parts.pop(-1)
        version = file_parts.pop(-1)
        etype = file_parts.pop(0)
        if (etype == "object" or etype == "collection" or etype == "dataset"):
            etype = etype + "_" + file_parts.pop(0)
        code = "_".join(file_parts)
            
        if not re.match(r"^(collection_type|object_type|dataset_type|vocabulary)$", etype):
            errors.append("<strong>Invalid entity type</strong> at position 1.")
        if not re.match(r"^([\w.]+)$", code):
            errors.append("<strong>Invalid entity name</strong> at position 2.")
        if not re.match(r"^(v\d+)$", version):
            errors.append("<strong>Invalid version</strong> at position 3.")
        if not re.match(r"^([a-zA-Z0-9]+(?:\.[0-9]+)?)$", section):
            errors.append("<strong>Invalid division</strong> at position 4.")
        if not re.match(r"^[a-zA-Z0-9]+$", creator):
            errors.append("<strong>Invalid contact person</strong> at position 5.")
            
        return ["\n".join(errors), False]
    

def index_to_excel_column(index):
    column = ''
    while index > 0:
        index, remainder = divmod(index - 1, 26)
        column = chr(65 + remainder) + column
    return column


def check_properties(sheet, errors):
    expected_terms = [
        "Version",
        "Code",
        "Description",
        "Mandatory",
        "Show in edit views",
        "Section",
        "Property label",
        "Data type",
        "Vocabulary code"
    ]
    row_headers = [cell.value for cell in sheet[4]]
    for term in expected_terms:
        if (term not in row_headers):
            if term in ("Mandatory","Show in edit views","Section"):
                errors.append(f"<em>Warning</em>: '{term}' not found in the properties headers.")
            else:
                errors.append(f"<strong>Error</strong>: '{term}' not found in the properties headers.")
        else:
             # Find the index of the term in the second row
             term_index = row_headers.index(term) + 1
             term_letter = index_to_excel_column(term_index)
             #print(term_index)
             
             # Check the column below "Version"
             if term == "Version":
                 column_below_version = []
                 for cell in sheet[term_letter][4:]:
                     if cell.value is not None:
                         column_below_version.append(cell.value)
                     else:
                         pass

                 # Check if any value in the column is not an integer
                 non_integer_indices = [i + 5 for i, cell in enumerate(column_below_version) if not (str(cell).isnumeric() or "$" in str(cell))]
                 if non_integer_indices:
                     # Append an error indicating the positions (row numbers) that are not integers
                     errors.append(f"<strong>Error</strong>: Values not valid found in the 'Version' column (they should be Integers) at row(s): {', '.join(map(str, non_integer_indices))}")

            # Check the column below "Code"
             elif term == "Code":
                column_below_code = []
                for cell in sheet[term_letter][4:]:
                    if cell.value is not None:
                        column_below_code.append(cell.value)
                    else:
                        pass
                invalid_codes = [i + 5 for i, cell in enumerate(column_below_code) if not (re.match(r'^\$?[A-Z0-9_.]+$', str(cell)) or "$" in str(cell))]
                if invalid_codes:
                    # Append an error indicating the positions (row numbers) with invalid values for the current term
                    errors.append(f"<strong>Error</strong>: Invalid code found in the '{term}' column at row(s): {', '.join(map(str, invalid_codes))}")
                    
                #check that all the properties of the object are different using a set (unique terms):
                if len(set(column_below_code)) != len(column_below_code):
                    seen_props = set()
                    repeated_props = set()
                    for prop in column_below_code:
                        if prop in seen_props:
                            repeated_props.add(prop)
                        else:
                            seen_props.add(prop)
                    errors.append(f"<strong>Error</strong>: The following properties are repeated: {repeated_props}. Please, delete the duplicates, and leave just one occurence")

            
            
            # Check the cell below "Description"
             elif term == "Description":
                column_below_description = []
                for cell in sheet[term_letter][4:]:
                    if cell.value is not None:
                        column_below_description.append(cell.value)
                    else:
                        pass
                invalid_indices = [i + 5 for i, cell in enumerate(column_below_description) if not (re.match(r'.*//.*', str(cell)) or "$" in str(cell))]
                if invalid_indices:
                    errors.append(f"<strong>Error</strong>: Invalid value(s) found in the '{term}' column at row(s): {', '.join(map(str, invalid_indices))}. Description should follow the schema: English Description + '//' + German Description.")

            # Check the cell below "Mandatory"
             elif term == "Mandatory":
                column_below_mandatory = []
                for cell in sheet[term_letter][4:]:
                    if cell.value is not None:
                        column_below_mandatory.append(str(cell.value).upper())
                    else:
                        pass
                invalid_mandatory = [i + 5 for i, cell in enumerate(column_below_mandatory) if (cell not in ["TRUE", "FALSE"] and "$" not in str(cell))]
                if invalid_mandatory:
                    errors.append(f"<strong>Error</strong>: Invalid value found in the '{term}' column at row(s): {', '.join(map(str, invalid_mandatory))}. Accepted values: TRUE, FALSE")

            # Check the cell below "Show in edit views"
             elif term == "Show in edit views":
                column_below_show = []
                for cell in sheet[term_letter][4:]:
                    if cell.value is not None:
                        column_below_show.append(str(cell.value).upper())
                    else:
                        pass
                invalid_show = [i + 5 for i, cell in enumerate(column_below_show) if (cell not in ["TRUE", "FALSE"] and "$" not in str(cell))]
                if invalid_show:
                    errors.append(f"<strong>Error</strong>: Invalid value found in the '{term}' column at row(s): {', '.join(map(str, invalid_show))}. Accepted values: TRUE, FALSE")

            # Check the cell below "Section"
             elif term == "Section":
                column_below_section = []
                for cell in sheet[term_letter][4:]:
                    if cell.value is not None:
                        column_below_section.append(cell.value) if '$' not in cell.value else column_below_section.append(cell.value.replace('$', ''))
                    else:
                        pass
                    
                invalid_section = [i + 5 for i, cell in enumerate(column_below_section) if not (re.match(r'^[A-Z][A-Za-z]*(?:\s[A-Z][A-Za-z]*)*$', str(cell)) or "$" in str(cell))]
                if invalid_section:
                    errors.append(f"<strong>Error</strong>: Invalid value found in the '{term}' column at row(s): {', '.join(map(str, invalid_section))}. Each word in the Section should start with a capital letter.")
            
                # Group Check: Ensure all properties within the same section are grouped together
                seen_sections = {}
                non_contiguous_rows = []
                
                for i, current_value in enumerate(column_below_section):
                    if current_value in seen_sections:
                        # If the value has been seen before but the row is not contiguous, add an error
                        if seen_sections[current_value] != i - 1:
                            non_contiguous_rows.append(i + 5)
                    seen_sections[current_value] = i  # Update the last seen row index for the current value
                
                if non_contiguous_rows:
                    errors.append(f"<strong>Error</strong>: Non-contiguous rows found for the same 'Section' value at row(s): {', '.join(map(str, non_contiguous_rows))}. Ensure that all properties within the same Section are grouped together.")
            
                # Predefined section order (fixed order)
                predefined_section_order = ["General Information", "Additional Information", "Comments"]
            
                # Validate contiguous groups and predefined section order
                seen_sections = set()
                previous_section_type = None
                section_errors = []  # Store section-specific errors
                additional_info_seen = False  # Flag to track if "Additional Information" has been seen
                comments_seen = False  # Flag to track if "Comments" has been seen
            
                # Traverse the section list
                for i, section in enumerate(column_below_section):
                    if section in predefined_section_order:
                        if section == "General Information":
                            if previous_section_type not in [None, "General Information"]:
                                section_errors.append(f"<strong>Error</strong> at row {i + 5}: 'General Information' should only appear at the beginning.")
                        elif section == "Additional Information":
                            if previous_section_type not in ["General Information", "user-defined"]:
                                section_errors.append(f"<strong>Error</strong> at row {i + 5}: 'Additional Information' should appear after 'General Information' and any user-defined sections.")
                            additional_info_seen = True  # Mark that "Additional Information" has been encountered
                        elif section == "Comments":
                            if previous_section_type not in ["General Information", "user-defined", "Additional Information"]:
                                section_errors.append(f"<strong>Error</strong> at row {i + 5}: 'Comments' should appear after 'Additional Information'.")
                            comments_seen = True  # Mark that "Comments" has been encountered
                        previous_section_type = section
                    else:
                        # User-defined section
                        if comments_seen:
                            section_errors.append(f"<strong>Error</strong> at row {i + 5}: User-defined section '{section}' cannot appear after 'Comments'.")
                        if additional_info_seen and not comments_seen:
                            section_errors.append(f"<strong>Error</strong> at row {i + 5}: User-defined section '{section}' cannot appear after 'Additional Information' but before 'Comments'.")
                        previous_section_type = "user-defined"
            
                # Output any errors
                if section_errors:
                    for error in section_errors:
                        errors.append(error)
            
            # Check the cell below "Property label"
             elif term == "Property label":
                column_below_label = []
                for cell in sheet[term_letter][4:]:
                    if cell.value is not None:
                        column_below_label.append(cell.value)
                    else:
                        pass
                invalid_label = [i + 5 for i, cell in enumerate(column_below_label) if not (re.match(r'.*', str(cell)) or "$" in str(cell))]
                if invalid_label:
                    errors.append(f"<strong>Error</strong>: Invalid value found in the '{term}' column at row(s): {', '.join(map(str, invalid_label))}. Specify the property label as text format")
                    
                # Dynamically find the "Section" column
                if "Section" in row_headers:
                    section_index = row_headers.index("Section") + 1
                    section_letter = index_to_excel_column(section_index)
                    column_below_section = [cell.value for cell in sheet[section_letter][4:]]

                    # New check: "Notes" in "Property label" should correspond to "Additional Information" in "Section"
                    for i, label_value in enumerate(column_below_label):
                        if label_value == "Notes":
                            section_value = column_below_section[i]
                            if section_value != "Additional Information":
                                errors.append(f"<strong>Error</strong>: 'Notes' found in the 'Property label' column at row {i + 5}, but corresponding 'Section' column does not contain 'Additional Information'. Value found: {section_value}")


            # Check the cell below "Data type"
             elif term == "Data type":
                column_below_type = []
                for cell in sheet[term_letter][4:]:
                    if cell.value is not None:
                        column_below_type.append(str(cell.value).upper())
                    else:
                        pass
                invalid_type = [i + 5 for i, cell in enumerate(column_below_type) if (cell not in ["INTEGER", "REAL", "VARCHAR", "MULTILINE_VARCHAR", "HYPERLINK", "BOOLEAN", "CONTROLLEDVOCABULARY", "XML", "TIMESTAMP", "DATE", "SAMPLE"] and "$" not in str(cell))]
                if invalid_type:
                    errors.append(f"<strong>Error</strong>: Invalid value found in the '{term}' column at row(s): {', '.join(map(str, invalid_type))}. Accepted types: INTEGER, REAL, VARCHAR, MULTILINE_VARCHAR, HYPERLINK, BOOLEAN, CONTROLLEDVOCABULARY, XML, TIMESTAMP, DATE, SAMPLE")

            # Check the column below "Vocabulary code"
             elif term == "Vocabulary code":
                column_below_vocab = sheet[term_letter][4:]
                invalid_vocab = [i + 5 for i, cell in enumerate(column_below_vocab) if cell.value and not (re.match(r'^\$?[A-Z0-9_.]', str(cell.value)) or "$" not in str(cell))]
                if invalid_vocab:
                    # Append an error indicating the positions (row numbers) with invalid values for the current term
                    errors.append(f"<strong>Error</strong>: Invalid vocabulary code found in the '{term}' column at row(s): {', '.join(map(str, invalid_vocab))}")
    
    return errors

def check_vocab_terms(sheet, errors):
    expected_terms = [
        "Version",
        "Code",
        "Label"
        "Description"
    ]
    row_headers = [cell.value for cell in sheet[4]]
    for term in expected_terms:
        if term not in row_headers:
            errors.append(f"<strong>Error</strong>: '{term}' not found in the vocabulary term headers.")
        else:
             # Find the index of the term in the second row
             term_index = row_headers.index(term) + 1
             term_letter = index_to_excel_column(term_index)
             #print(term_index)
             
             # Check the column below "Version"
             if term == "Version":
                 column_below_version = []
                 for cell in sheet[term_letter][4:]:
                     if cell.value is not None:
                         column_below_version.append(cell.value)
                     else:
                         pass

                 # Check if any value in the column is not an integer
                 non_integer_indices = [i + 5 for i, cell in enumerate(column_below_version) if not str(cell).isnumeric()]
                 if non_integer_indices:
                     # Append an error indicating the positions (row numbers) that are not integers
                     errors.append(f"<strong>Error</strong>: Values not valid found in the 'Version' column (they should be Integers) at row(s): {', '.join(map(str, non_integer_indices))}")

            # Check the column below "Code"
             elif term == "Code":
                column_below_code = []
                for cell in sheet[term_letter][4:]:
                    if cell.value is not None:
                        column_below_code.append(cell.value)
                    else:
                        pass
                invalid_codes = [i + 5 for i, cell in enumerate(column_below_code) if not re.match(r'^\$?[A-Z0-9_.]+$', str(cell))]
                if invalid_codes:
                    # Append an error indicating the positions (row numbers) with invalid values for the current term
                    errors.append(f"<strong>Error</strong>: Invalid code found in the '{term}' column at row(s): {', '.join(map(str, invalid_codes))}")
                
                #check that all the properties of the object are different using a set (unique terms):
                if len(set(column_below_code)) != len(column_below_code):
                    seen_terms = set()
                    repeated_terms = set()
                    for term in column_below_code:
                        if term in seen_terms:
                            repeated_terms.add(term)
                        else:
                            seen_terms.add(term)
                    errors.append(f"<strong>Error</strong>: The following vocabulary terms are repeated: {repeated_terms}. Please, delete the duplicates, and leave just one occurence")

            
            
            # Check the cell below "Description"
             elif term == "Description":
                column_below_description = sheet[term_letter][4:]
                invalid_description = [i + 5 for i, cell in enumerate(column_below_description) if cell.value and not re.match(r'.*//.*', str(cell.value))]
                if invalid_description:
                    errors.append(f"<strong>Error</strong>: Invalid value(s) found in the '{term}' column at row(s): {', '.join(map(str, invalid_description))}. Description should follow the schema: English Description + '//' + German Description.")

            # Check the cell below "Mandatory"
             elif term == "Label":
                column_below_label = sheet[term_letter][4:]
                invalid_label = [i + 5 for i, cell in enumerate(column_below_label) if cell.value and not re.match(r'.*', str(cell.value))]
                if invalid_label:
                    errors.append(f"<strong>Error</strong>: Invalid value found in the '{term}' column at row(s): {', '.join(map(str, invalid_label))}. Specify the label as text format")
            
    return "\n".join(errors)

def content_checker(file_name, name_ok):
    logger.info(f"Type {type(file_name)} of file {file_name}")
    workbook = openpyxl.load_workbook(file_name)
    errors = []  
    
    if(name_ok):
        file_name = file_name.name.split(".xls")
        file_parts = file_name[0].split("_")
        file_parts.pop(-1)
        file_parts.pop(-1)
        version = file_parts.pop(-1)
        etype = file_parts.pop(0)
        if (etype == "object" or etype == "collection" or etype == "dataset"):
            etype = etype + "_" + file_parts.pop(0)
        code = "_".join(file_parts)
    else:
        version, etype, code = "", "", ""

    sheet = workbook.active
    
    filtered_rows = []
    
    for row in sheet.iter_rows(min_row=1, values_only=True):
    # Check if any cell in the row contains "$"
        if any("$" in str(cell) for cell in row):
            filtered_rows.append(["$" + str(cell) if cell is not None else None for cell in row])
        else:
            # If the row passed the check, add it to the filtered list
            filtered_rows.append(row)
    
    #remove all the rows in the sheet
    sheet.delete_rows(0, sheet.max_row)

    # Append the filtered rows to the sheet
    for row_data in filtered_rows:
        sheet.append(row_data)

    # Access a specific cell (e.g., cell A1)
    cell_value_A1 = sheet['A1'].value
    print(f"Entity Type: {cell_value_A1}")
    
    entity_types = ["SAMPLE_TYPE", "EXPERIMENT_TYPE", "DATASET_TYPE", "PROPERTY_TYPE", "VOCABULARY_TYPE"]
    if cell_value_A1 not in entity_types:
        errors.append("The entity type (cell A1) should be one of the following: SAMPLE_TYPE, EXPERIMENT_TYPE, DATASET_TYPE, PROPERTY_TYPE, VOCABULARY_TYPE")
        return "".join(errors)
    else:
        if cell_value_A1 == "SAMPLE_TYPE":
            expected_terms = [
                "Version",
                "Code",
                "Description",
                "Validation script",
                "Generated code prefix",
                "Auto generate codes",
            ]
            second_row_values = [cell.value for cell in sheet[2]]
            for term in expected_terms:
                if term not in second_row_values:
                    errors.append(f"<strong>Error</strong>: '{term}' not found in the entity headers.")
                else:
                     # Find the index of the term in the second row
                     term_index = second_row_values.index(term)

                     # Check the cell below "Version"
                     if term == "Version":
                        cell_below_version = sheet.cell(row=3, column=term_index + 1)
                        if str(cell_below_version.value) != version[1:]:
                            errors.append("<strong>Error</strong>: The version should be the same one indicated in the file name")

                    # Check the cell below "Code"
                     elif term == "Code":
                        cell_below_code = sheet.cell(row=3, column=term_index + 1)
                        if cell_below_code.value != code:
                            errors.append("<strong>Error</strong>: The code should be the same one indicated in the file name")
                    
                    
                    # Check the cell below "Description"
                     elif term == "Description":
                        cell_below_description = sheet.cell(row=3, column=term_index + 1)
                        description_pattern = re.compile(r".*//.*")
                        if not description_pattern.match(cell_below_description.value):
                            errors.append("<strong>Error</strong>: Description should follow the schema: English Description + '//' + German Description.")

                    # Check the cell below "Generated code prefix"
                     elif term == "Generated code prefix":
                        cell_below_generated_code = sheet.cell(row=3, column=term_index + 1)
                        code_replace = code.replace('_', '.').split('.')
                        ext_code = [word[:3].upper() for word in code_replace]
                        generated_code = '.'.join(ext_code)
                        if cell_below_generated_code.value != generated_code:
                            errors.append("<em>Warning</em>: It is recommended that the value of 'Generated code prefix' be the first three letters of each part of the 'Code' separated by dots ['.'].")

                    # Check the cell below "Validation script"
                     elif term == "Validation script":
                        cell_below_validation = sheet.cell(row=3, column=term_index + 1)
                        validation_pattern = re.compile(r"^[A-Za-z0-9_]+\.py$")
                        if cell_below_validation.value and not validation_pattern.match(cell_below_validation.value):
                             errors.append("<strong>Error</strong>: Validation script should follow the schema: Words and/or numbers separated by '_' and ending in '.py'")


                    # Check the cell below "Auto generate codes"
                     elif term == "Auto generate codes":
                        cell_below_auto_generate = sheet.cell(row=3, column=term_index + 1)
                        auto_code = cell_below_auto_generate.value
                        if (auto_code == True): auto_code = "TRUE"
                        if (auto_code == False): auto_code = "FALSE"
                        if auto_code not in ["TRUE", "FALSE"]:
                            errors.append("<strong>Error</strong>: Value below 'Auto generate codes' should be 'TRUE' or 'FALSE'.")
            
            errors = check_properties(sheet, errors)      
            
        elif cell_value_A1 == "EXPERIMENT_TYPE" or cell_value_A1 == "DATASET_TYPE":
            expected_terms = [
                "Version",
                "Code",
                "Description",
                "Validation script"
            ]
            second_row_values = [cell.value for cell in sheet[2]]
            for term in expected_terms:
                if term not in second_row_values:
                    errors.append(f"<strong>Error</strong>: '{term}' not found in the second row.")
                else:
                     # Find the index of the term in the second row
                     term_index = second_row_values.index(term)

                     # Check the cell below "Version"
                     if term == "Version":
                        cell_below_version = sheet.cell(row=3, column=term_index + 1)
                        if str(cell_below_version.value) != version[1:]:
                            errors.append("<strong>Error</strong>: The version should be the same one indicated in the file name")

                    # Check the cell below "Code"
                     elif term == "Code":
                        cell_below_code = sheet.cell(row=3, column=term_index + 1)
                        if cell_below_code.value != code:
                            errors.append("<strong>Error</strong>: The code should be the same one indicated in the file name")
                    
                    
                    # Check the cell below "Description"
                     elif term == "Description":
                        cell_below_description = sheet.cell(row=3, column=term_index + 1)
                        description_pattern = re.compile(r".*//.*")
                        if not description_pattern.match(cell_below_description.value):
                            errors.append("<strong>Error</strong>: Description should follow the schema: English Description + '//' + German Description.")
            
            
                    # Check the cell below "Validation script"
                     elif term == "Validation script":
                        cell_below_validation = sheet.cell(row=3, column=term_index + 1)
                        validation_pattern = re.compile(r"^[A-Za-z0-9_]+\.py$")
                        if cell_below_validation.value and not validation_pattern.match(cell_below_validation.value):
                            errors.append("<strong>Error</strong>: Validation script should follow the schema: Words and/or numbers separated by '_' and ending in '.py'")

            errors = check_properties(sheet, errors) 
            
        elif cell_value_A1 == "VOCABULARY_TYPE":
            expected_terms = [
                "Version",
                "Code",
                "Description"
            ]
            second_row_values = [cell.value for cell in sheet[2]]
            for term in expected_terms:
                if term not in second_row_values:
                    errors.append(f"<strong>Error</strong>: '{term}' not found in the second row.")
                else:
                     # Find the index of the term in the second row
                     term_index = second_row_values.index(term)

                     # Check the cell below "Version"
                     if term == "Version":
                        cell_below_version = sheet.cell(row=3, column=term_index + 1)
                        if str(cell_below_version.value) != version[1:]:
                            errors.append("<strong>Error</strong>: The version should be the same one indicated in the file name. Value found: {cell_below_version.value}")

                    # Check the cell below "Code"
                     elif term == "Code":
                        cell_below_code = sheet.cell(row=3, column=term_index + 1)
                        if cell_below_code.value != code:
                            errors.append("<strong>Error</strong>: The code should be the same one indicated in the file name. Value found: {cell_below_code.value}")
                    
                    
                    # Check the cell below "Description"
                     elif term == "Description":
                        cell_below_description = sheet.cell(row=3, column=term_index + 1)
                        description_pattern = re.compile(r".*//.*")
                        if not description_pattern.match(cell_below_description.value):
                            errors.append("<strong>Error</strong>: Description should follow the schema: English Description + '//' + German Description. Value found: {cell_below_description.value}")
            
            errors = check_vocab_terms(sheet, errors)

        elif cell_value_A1 == "PROPERTY_TYPE":
            expected_terms = [
                "Version",
                "Code",
                "Description",
                "Mandatory",
                "Show in edit views",
                "Section",
                "Property label",
                "Data type",
                "Vocabulary code"
            ]
            second_row_values = [cell.value for cell in sheet[2]]
            for term in expected_terms:
                if term not in second_row_values:
                    errors.append(f"<strong>Error</strong>: '{term}' not found in the second row.")
                else:
                     # Find the index of the term in the second row
                     term_index = second_row_values.index(term) + 1


                     # Check the column below "Version"
                     if term == "Version":
                        column_below_version = sheet[term_index][2:]
                        # Check if any value in the column is not an integer
                        non_integer_cells = [(i + 3, cell.value) for i, cell in enumerate(column_below_version) if not isinstance(cell.value, int)]
                        if non_integer_cells:
                            # Append an error indicating the positions (row numbers) that are not integers
                            non_integer_indices = [str(row) for row, _ in non_integer_cells]
                            invalid_values = [str(value) for _, value in non_integer_cells]
                            errors.append(f"<strong>Error</strong>: Values not valid found in the 'Version' column (they should be Integers) at row(s): {', '.join(non_integer_indices)}. Value(s) found: {', '.join(invalid_values)}")

                    # Check the column below "Code"
                     elif term == "Code":
                        column_below_code = sheet[term_index][2:]
                        invalid_codes = [(i + 3, cell.value) for i, cell in enumerate(column_below_code) if not re.match(r'^\$?[A-Z0-9_.]+$', str(cell.value))]
                        if invalid_codes:
                            invalid_rows = [str(row) for row, _ in invalid_codes]
                            invalid_values = [str(value) for _, value in invalid_codes]
                            errors.append(f"<strong>Error</strong>: Invalid code found in the '{term}' column at row(s): {', '.join(invalid_rows)}. Value(s) found: {', '.join(invalid_values)}")
                    
                    # Check the cell below "Description"
                     elif term == "Description":
                        column_below_description = sheet[term_index][2:]
                        invalid_descriptions = [(i + 3, cell.value) for i, cell in enumerate(column_below_description) if not re.match(r'.*//.*', str(cell.value))]
                        if invalid_descriptions:
                            invalid_rows = [str(row) for row, _ in invalid_descriptions]
                            invalid_values = [str(value) for _, value in invalid_descriptions]
                            errors.append(f"<strong>Error</strong>: Invalid value(s) found in the '{term}' column at row(s): {', '.join(invalid_rows)}. Description should follow the schema: English Description + '//' + German Description. Value(s) found: {', '.join(invalid_values)}")
                    
                    # Check the cell below "Mandatory"
                     elif term == "Mandatory":
                        column_below_mandatory = sheet[term_index][2:]
                        invalid_mandatory = [(i + 3, cell.value) for i, cell in enumerate(column_below_mandatory) if cell.value not in ["TRUE", "FALSE"]]
                        if invalid_mandatory:
                            invalid_rows = [str(row) for row, _ in invalid_mandatory]
                            invalid_values = [str(value) for _, value in invalid_mandatory]
                            errors.append(f"<strong>Error</strong>: Invalid value found in the '{term}' column at row(s): {', '.join(invalid_rows)}. Accepted values: TRUE, FALSE. Value(s) found: {', '.join(invalid_values)}")
                    
                    # Check the cell below "Show in edit views"
                     elif term == "Show in edit views":
                        column_below_show = sheet[term_index][2:]
                        invalid_show = [(i + 3, cell.value) for i, cell in enumerate(column_below_show) if cell.value not in ["TRUE", "FALSE"]]
                        if invalid_show:
                            invalid_rows = [str(row) for row, _ in invalid_show]
                            invalid_values = [str(value) for _, value in invalid_show]
                            errors.append(f"<strong>Error</strong>: Invalid value found in the '{term}' column at row(s): {', '.join(invalid_rows)}. Accepted values: TRUE, FALSE. Value(s) found: {', '.join(invalid_values)}")
                    
                     elif term == "Section":
                        column_below_section = sheet[term_index][2:]
                        print(column_below_section)
                        invalid_section = [(i + 3, cell.value) for i, cell in enumerate(column_below_section) if not re.match(r'^[A-Z][a-z]*(?:\s[A-Z][a-z]*)*$', str(cell.value))]
                        if invalid_section:
                            invalid_rows = [str(row) for row, _ in invalid_section]
                            invalid_values = [str(value) for _, value in invalid_section]
                            errors.append(f"<strong>Error</strong>: Invalid value found in the '{term}' column at row(s): {', '.join(invalid_rows)}. Each word should start with a capital letter. Value(s) found: {', '.join(invalid_values)}")
                    
                        seen_sections = {}
                        non_contiguous_rows = []

                        for i, current_value in enumerate(column_below_section):
                            if current_value in seen_sections:
                                # If the value has been seen before but the row is not contiguous, add an error
                                if seen_sections[current_value] != i - 1:
                                    non_contiguous_rows.append(i + 5)
                            seen_sections[current_value] = i  # Update the last seen row index for the current value

                        if non_contiguous_rows:
                            errors.append(f"<strong>Error</strong>: Non-contiguous rows found for the same 'Section' value at row(s): {', '.join(map(str, non_contiguous_rows))}. Ensure that all properties within the same Section are grouped together.")
                            
                        # Predefined section order (fixed order)
                        predefined_section_order = ["General Information", "Additional Information", "Comments"]
                    
                        # Validate contiguous groups and predefined section order
                        seen_sections = set()
                        previous_section_type = None
                        section_errors = []  # Store section-specific errors
                        additional_info_seen = False  # Flag to track if "Additional Information" has been seen
                        comments_seen = False  # Flag to track if "Comments" has been seen
                    
                        # Traverse the section list
                        for i, section in enumerate(column_below_section):
                            if section in predefined_section_order:
                                if section == "General Information":
                                    if previous_section_type not in [None, "General Information"]:
                                        section_errors.append(f"<strong>Error</strong> at row {i + 5}: 'General Information' should only appear at the beginning.")
                                elif section == "Additional Information":
                                    if previous_section_type not in ["General Information", "user-defined"]:
                                        section_errors.append(f"<strong>Error</strong> at row {i + 5}: 'Additional Information' should appear after 'General Information' and any user-defined sections.")
                                    additional_info_seen = True  # Mark that "Additional Information" has been encountered
                                elif section == "Comments":
                                    if previous_section_type not in ["General Information", "user-defined", "Additional Information"]:
                                        section_errors.append(f"<strong>Error</strong> at row {i + 5}: 'Comments' should appear after 'Additional Information'.")
                                    comments_seen = True  # Mark that "Comments" has been encountered
                                previous_section_type = section
                            else:
                                # User-defined section
                                if comments_seen:
                                    section_errors.append(f"<strong>Error</strong> at row {i + 5}: User-defined section '{section}' cannot appear after 'Comments'.")
                                if additional_info_seen and not comments_seen:
                                    section_errors.append(f"<strong>Error</strong> at row {i + 5}: User-defined section '{section}' cannot appear after 'Additional Information' but before 'Comments'.")
                                previous_section_type = "user-defined"
                    
                        # Output any errors
                        if section_errors:
                            for error in section_errors:
                                errors.append(error)
            
                    # Check the cell below "Property label"
                     elif term == "Property label":
                        column_below_label = sheet[term_index][2:]
                        invalid_label = [(i + 3, cell.value) for i, cell in enumerate(column_below_label) if not re.match(r'.*', str(cell.value))]
                        if invalid_label:
                            invalid_rows = [str(row) for row, _ in invalid_label]
                            invalid_values = [str(value) for _, value in invalid_label]
                            errors.append(f"<strong>Error</strong>: Invalid value found in the '{term}' column at row(s): {', '.join(invalid_rows)}. Specify the property label as text format. Value(s) found: {', '.join(invalid_values)}")
                         # Dynamically find the "Section" column
                        if "Section" in second_row_values:
                            section_column_index = second_row_values.index("Section") + 1  # Find the index of the "Section" column
                            section_letter = index_to_excel_column(section_column_index)  # Convert index to Excel column letter
                            column_below_section = sheet[section_letter][2:]  # Get all cells below the "Section" header

                            # New check for "Notes" in "Property label" and "Additional Information" in "Section"
                            for i, cell in enumerate(column_below_label):
                                if cell.value == "Notes":
                                    section_value = column_below_section[i].value  # Get the value in the "Section" column for the same row
                                    if section_value != "Additional Information":
                                        errors.append(f"<strong>Error</strong>: 'Notes' found in the 'Property label' column at row {i + 5}, but corresponding 'Section' column does not contain 'Additional Information'. Value found: {section_value}")
                    
                    # Check the cell below "Data type"
                     elif term == "Data type":
                        column_below_type = sheet[term_index][2:]
                        invalid_type = [(i + 3, cell.value) for i, cell in enumerate(column_below_type) if cell.value not in ["INTEGER", "REAL", "VARCHAR", "MULTILINE_VARCHAR", "HYPERLINK", "BOOLEAN", "CONTROLLEDVOCABULARY", "XML", "TIMESTAMP", "DATE", "SAMPLE"]]
                        if invalid_type:
                            invalid_rows = [str(row) for row, _ in invalid_type]
                            invalid_values = [str(value) for _, value in invalid_type]
                            errors.append(f"<strong>Error</strong>: Invalid value found in the '{term}' column at row(s): {', '.join(invalid_rows)}. Accepted types: INTEGER, REAL, VARCHAR, MULTILINE_VARCHAR, HYPERLINK, BOOLEAN, CONTROLLEDVOCABULARY, XML, TIMESTAMP, DATE, SAMPLE. Value(s) found: {', '.join(invalid_values)}")
                    
                    # Check the column below "Vocabulary code"
                     elif term == "Vocabulary code":
                        column_below_vocab = sheet[term_index][2:]
                        invalid_vocab = [(i + 3, cell.value) for i, cell in enumerate(column_below_vocab) if cell.value is not None and not re.match(r'^\$?[A-Z0-9_.]+$', str(cell.value))]
                        if invalid_vocab:
                            invalid_rows = [str(row) for row, _ in invalid_vocab]
                            invalid_values = [str(value) for _, value in invalid_vocab]
                            errors.append(f"<strong>Error</strong>: Invalid vocabulary code found in the '{term}' column at row(s): {', '.join(invalid_rows)}. Value(s) found: {', '.join(invalid_values)}")


    # Close the workbook after use
    workbook.close()
    print(type(errors))
    print(errors)
    if type(errors) == list:
        output = "\n".join(errors)
    else:
        output = "".join(errors)
    if output == "":
        return "File content: OK!"
    else:
        return output